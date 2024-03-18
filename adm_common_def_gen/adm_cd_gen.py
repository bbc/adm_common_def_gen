#!/usr/bin/python
"""
Generates a Standard Definitions ADM XML file from a spreadsheet containing channel and pack information
"""

from sys import *
import xlrd
from .adm_write_xml import *
from copy import deepcopy
from openpyxl import load_workbook

def ReadXLS(fname, sheet_name):
    """
    Read an XLS file. Assumes first sheet with third row as headers
    """
    
    FIRST_INFO_ROW = 3
    
    book = load_workbook(filename = fname)
    sheet = book[sheet_name]

    # Find the ID header cell position (i_pos) which splits the channels (to the left) and pack info (to the right)
    i_pos = 0
    for c in range(1, sheet.max_column):
        if sheet.cell(row=1, column=c).value == "ID":
            i_pos = c
            
    # Get the labels for the channel info from the headers
    ch_labels = []
    for col in range(1, i_pos):
        ch_labels.append(sheet.cell(row=FIRST_INFO_ROW, column=col).value)

    # Get all the channel definitions into all_ch
    all_ch = []
    for row in range(FIRST_INFO_ROW+1, sheet.max_row + 1):
        # Use labels as dict keys for entries
        ch_data = {}
        for col in range(1, i_pos):
            ch_data[ch_labels[col - 1]] = sheet.cell(row=row, column=col).value
        all_ch.append(ch_data)

    # Get the pack names and IDs with an index for each
    pa_names = []
    i = 0
    for col in range(i_pos + 1, sheet.max_column + 1):
        pa_names.append({'ind': i, 
                         'name': sheet.cell(row=FIRST_INFO_ROW, column=col).value, 
                         'id': sheet.cell(row=1, column=col).value, 
                         'urn': sheet.cell(row=2, column=col).value})
        i += 1

    # Read each pack column and assign channel IDs according to their number in the table
    i = 0
    pa_list = []
    for col in range(i_pos + 1, sheet.max_column + 1):
        co = 1
        ch_list = []
        for goes in range(FIRST_INFO_ROW + 1, sheet.max_row + 1):
            for row in range(FIRST_INFO_ROW + 1, sheet.max_row + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    if int(float(sheet.cell(row=row, column=col).value)) == co:
                        ch_list.append(all_ch[row - FIRST_INFO_ROW - 1]['ID'])
                        co += 1
        pa_list.append({'ID': pa_names[i], 'list': ch_list})
        i += 1
    
    return all_ch, pa_list  

  
def ReadXLSHOA(fname):
    """
    Read an XLS file. Assumes second sheet with third row as headers
    """
    
    FIRST_INFO_ROW = 3
    
    book = load_workbook(filename = fname)
    sheet = book['HOA']

    # Find the ID header cell position (i_pos) which splits the channels (to the left) and pack info (to the right)
    i_pos = 0
    h_pos = 0
    for c in range(1, sheet.max_column):
        if sheet.cell(row=1, column=c).value == "ID":
            i_pos = c
        if (sheet.cell(row=FIRST_INFO_ROW, column=c).value is None) and (h_pos == 0):
            h_pos = c

    # Get the labels for the channel info from the headers
    ch_labels = []
    for col in range(1, h_pos):
        if sheet.cell(row=FIRST_INFO_ROW, column=col).value is not None:
            ch_labels.append(sheet.cell(row=FIRST_INFO_ROW, column=col).value)

    # Get all the channel definitions into all_ch
    all_ch = []
    for row in range(FIRST_INFO_ROW+1, sheet.max_row + 1):
        # Use labels as dict keys for entries
        if sheet.cell(row=row, column=1).value is not None:
            ch_data = {}
            for col in range(1, h_pos):
                ch_data[ch_labels[col - 1]] = sheet.cell(row=row, column=col).value
            all_ch.append(ch_data)
        else:
            all_ch.append(None)

    pa_names = []
    i = 0
    for col in range(i_pos + 1, sheet.max_column + 1):
        if sheet.cell(row=FIRST_INFO_ROW, column=col).value is not None:
            pa_names.append({'ind': i, 
                            'name': sheet.cell(row=FIRST_INFO_ROW, column=col).value, 
                            'id': sheet.cell(row=1, column=col).value, 
                            'child': sheet.cell(row=2, column=col).value})
            i += 1

    i = 0
    pa_list = []
    for col in range(i_pos + 1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value is not None:
            co = 1
            ch_list = []
            for goes in range(FIRST_INFO_ROW + 1, sheet.max_row + 1):
                for row in range(FIRST_INFO_ROW + 1, sheet.max_row + 1):
                    if sheet.cell(row=row, column=col).value is not None:
                        if int(float(sheet.cell(row=row, column=col).value)) == co:
                            if all_ch[row - FIRST_INFO_ROW - 1] is not None:
                                ch_list.append(all_ch[row - FIRST_INFO_ROW - 1]['ID'])
                                co += 1
            pa_list.append({'ID': pa_names[i], 'list': ch_list})
            i += 1

    return all_ch, pa_list


def GenerateXMLPacks(admxml, pa_list, pa_list_hoa):
    """
    audioPackFormat generation
    """
    for pa in pa_list:
        p_id = pa['ID']['id']
        p_name = ''
        if pa['ID']['urn']:
          p_name += pa['ID']['urn']
        p_name += pa['ID']['name']
        chan_list = pa['list']
        if p_id[6] == '5':    # Binaural
            admxml.SetAudioPackFormat(p_id, p_name, chan_list, None, "0005", "Binaural")
        else:
            admxml.SetAudioPackFormat(p_id, p_name, chan_list, None, "0001", "DirectSpeakers")

    for pa in pa_list_hoa:
        p_id = pa['ID']['id']
        p_name = pa['ID']['name']
        chan_list = pa['list']
        if pa['ID']['child']:
            pack_list = [pa['ID']['child']]
        else:
            pack_list = []
        if p_id:
            admxml.SetAudioPackFormat(p_id, p_name, chan_list, pack_list, "0004", "HOA")
        
        
def GenerateXMLChannels(admxml, all_ch, all_ch_hoa):
    """
    audioChannelFormat generation
    """
    for ch in all_ch:
        if ch['ID'][6] == '5':    # Binaural
            block_list = []
            admxml.SetAudioChannelFormat(ch['ID'], ch['Name'], block_list, "0005", "Binaural")
        else:                    # DirectSpeakers
            pos = []
            edge = None
            if ch['Edge']:
                edge = ch['Edge']
            freq = None
            if ch['Frequency']:
                freq = ch['Frequency']
            if 'Azimuth' in ch:
                pos.append({'coordinate': 'azimuth', 'value': ch['Azimuth'], 'screenEdgeLock': edge})
                pos.append({'coordinate': 'elevation', 'value': ch['Elevation']})
                pos.append({'coordinate': 'distance', 'value': 1.0})
            elif 'X' in ch:
                pos.append({'coordinate': 'X', 'value': ch['X'], 'screenEdgeLock': edge})
                pos.append({'coordinate': 'Y', 'value': ch['Y']})
                pos.append({'coordinate': 'Z', 'value': ch['Z']})
            speaker_label = ""
            if ch['URI']:
                speaker_label += ch['URI']
            speaker_label += ch['Speaker']
            block_id = 'AB_' + ch['ID'][3:11] + '_00000001'
            block = {'id': block_id, 'speakerLabel': speaker_label, 'position': pos}
            block_list = [block]
            if freq:
                admxml.SetAudioChannelFormat(ch['ID'], ch['Name'], block_list, "0001", "DirectSpeakers", frequency_l=freq)
            else:
                admxml.SetAudioChannelFormat(ch['ID'], ch['Name'], block_list, "0001", "DirectSpeakers")
          
    for ch in all_ch_hoa:
        if ch:
            order = int(ch['Order'])
            degree = int(ch['Degree'])
            normalization = ch['Normalization']
            block_id = 'AB_' + ch['ID'][3:11] + '_00000001'
            block = {'id': block_id, 'order': order, 'degree': degree, 'normalization': normalization}
            block_list = [block]
            admxml.SetAudioChannelFormat(ch['ID'], ch['Name'], block_list, "0004", "HOA")
        

def GenerateXMLStreams(admxml, all_ch, all_ch_hoa):
    """
    audioStreamFormat generation
    """
    for ch in all_ch:
        s_id = "AS_%s" % (ch['ID'][3:])
        s_name = "PCM_%s" % (ch['Name'])
        track_list = ["AT_%s_01" % (ch['ID'][3:])]
        channel = ch['ID']
        admxml.SetAudioStreamFormat(s_id, s_name, track_list, channel, None, "0001", "PCM")
        
    # audioStreamFormat generation
    for ch in all_ch_hoa:
        if ch:
            s_id = "AS_%s" % (ch['ID'][3:])
            s_name = "PCM_%s" % (ch['Name'])
            track_list = ["AT_%s_01" % (ch['ID'][3:])]
            channel = ch['ID']
            admxml.SetAudioStreamFormat(s_id, s_name, track_list, channel, None, "0001", "PCM")
        
        
def GenerateXMLTracks(admxml, all_ch, all_ch_hoa):
    """
    audioTrackFormat generation
    """
    for ch in all_ch:
        t_id = "AT_%s_01" % (ch['ID'][3:])
        t_name = "PCM_%s" % (ch['Name'])
        stream = "AS_%s" % (ch['ID'][3:])
        admxml.SetAudioTrackFormat(t_id, t_name, stream, "0001", "PCM")

    # audioTrackFormat generation 
    for ch in all_ch_hoa:
        if ch:
            t_id = "AT_%s_01" % (ch['ID'][3:])
            t_name = "PCM_%s" % (ch['Name'])
            stream = "AS_%s" % (ch['ID'][3:])
            admxml.SetAudioTrackFormat(t_id, t_name, stream, "0001", "PCM")
 
 
def main():
    if len(argv) != 3:
        print("adm_cd_gen <input xls file> <output xml file>")
        return 0

    fout = open(argv[2], 'w')

    # List of all takes in sheet for channel-based
    all_ch, pa_list = ReadXLS(argv[1], 'Channel')
    all_ch_c, pa_list_c = ReadXLS(argv[1], 'ChannelCart')
    
    # Merge the polar and cartesian channels and packs
    all_ch.extend(all_ch_c)
    pa_list.extend(pa_list_c)

    # List of all takes in sheet for HOA-based
    all_ch_hoa, pa_list_hoa = ReadXLSHOA(argv[1])
    
    admxml = admXML()

    GenerateXMLPacks(admxml, pa_list, pa_list_hoa)
    GenerateXMLChannels(admxml, all_ch, all_ch_hoa)
    GenerateXMLStreams(admxml, all_ch, all_ch_hoa)
    GenerateXMLTracks(admxml, all_ch, all_ch_hoa)
    
    admxml.Write(fout)

    fout.close()
    
if __name__ == '__main__':
    main()
